#!/usr/bin/env python

'''
    2017-5-24
    FOE
    find out the elapse
    
    
    '''

from sc import *

import re
def isMatch(reEx, maStr):
    rst = re.match(reEx, maStr)
    if rst==None:
        return False
    else:
        return True

def isRegularExpressionWithDateTime(testString):
    if cmp(testString[:6],"2017-0") == 0:
        return True
    else:
        return False

import time
import datetime
def secondsFromDateTimeString(timeString):
    timeFormat = "%Y-%m-%d %H:%M:%S:%f"
    return datetime.datetime.strptime(timeString, timeFormat)
#    return time.mktime(time.strptime(timeString, timeFormat))

def integrateOne(lines,marks, dic):
    startline_ns=[-1]*len(marks)
    
    # search mark with format dic
    subs = []
    for sd in marks:
        sln = -1; # start line number
        filer_min_seconds = 0.00001
        if sd.has_key('filer_min_seconds'):
            filer_min_seconds = sd['filer_min_seconds']
        for i in range(len(lines)):
            curline = lines[i]
#            ind_s = isMatch(sd['mark_start'], curline)
            if isMatch(sd['mark_start'], curline):
                sln=i
#            ind_e = isMatch(sd['mark_end'], curline)
            if sln>0 and isMatch(sd['mark_end'], curline):
                cltime = secondsFromDateTimeString(str(curline[0:23]))
                sltime = secondsFromDateTimeString(str(lines[sln][0:23]))
#                print str(curline[0:23])
#                return
                elapse=(cltime-sltime).total_seconds()
                if filer_min_seconds > elapse:
                    continue
                tmp = {}
                tmp['elapse']=elapse
                tmp['start_line']=lines[sln].strip(' \n')
                tmp['end_line']=curline.strip(' \n')
                if sd.has_key('sub_marks'):
                    subm = sd['sub_marks']
                    integrateOne(lines[sln:i], subm, tmp)
                subs.append(tmp)
                start_line_number=-1
                
            pass
    dic['subs']=subs
    
    pass

import json
from openpyxl import Workbook
from openpyxl.styles import Alignment
def findOutElapse(fn, mi=0, out_par=0):
    ''' 
        file name
        
        '''
    
    mark_starts=[]
    mark_ends=[]
    
    mark_starts.append(r'  INFO server.mm run ')
    mark_ends.append(r'INFO OTFProcessor.mm run: 665 - Test case execute complete.')
    
    configPath = 'configure.json'
    if not os.path.exists(configPath):
        tmpfile=file('configure.json', 'w')
        defaultStr=r'''{
    "mark_start":".*INFO server\\.mm run 132 -.*",
    "mark_end":".*INFO OTFProcessor\\.mm run: 665 - Test case execute complete.*",
    "sub_marks":[
                    {
                    "mark_start":".*INFO libWrite_DUT_CB\\.m ReadAndWriteBin: 32 - Send --> 'cbwrite 0xD2 pass'.*",
                    "mark_end":".*INFO libWrite_DUT_CB\\.m executePlugin:andSN: 219 - Write CB 'pass' success.*"
                    }
                ]
    "filer_min_seconds":0.00001
}'''
        tmpfile.write(defaultStr)
        tmpfile.close()
    search_format = json.load(file(configPath, 'r'))
    
    if not os.path.exists(fn):
        print('log file not exist')
        return

    loglines = file(fn, 'r').readlines()
    filer_min_seconds = 0.00001
    if search_format.has_key('filer_min_seconds'):
        filer_min_seconds = search_format['filer_min_seconds']
    start_line_number=-1

    outBuff=[]

    print('lines count:%d'%(len(loglines)))
    for i in range(len(loglines)):
#        print i
        curline = loglines[i];
#        ind_s = isMatch(search_format['mark_start'], curline)
#        print 'm:%s ,cur:%s'%(search_format['mark_start'], curline)
        if isMatch(search_format['mark_start'], curline):
#            print 'start line', curline
            start_line_number=i
#        ind_e = curline.find(search_format['mark_end'])
        if start_line_number>0 and isMatch(search_format['mark_end'], curline):
#            print('find match: %s'%curline)
            cltime = secondsFromDateTimeString(str(curline[0:23]))
            sltime = secondsFromDateTimeString(str(loglines[start_line_number][0:23]))
            tmp = {}
            elapse=(cltime-sltime).total_seconds()
            if filer_min_seconds > elapse:
                continue
            tmp['elapse']=elapse
            tmp['start_line']=loglines[start_line_number].strip(' \n')
            tmp['end_line']=curline.strip(' \n')
            
            if search_format.has_key('sub_marks'):
                subm = search_format['sub_marks']
                integrateOne(loglines[start_line_number:i], subm, tmp)
            outBuff.append(tmp)
            start_line_number=-1

#        return
#    print('out:\n%s'%outBuff)
#    print json.dumps(outBuff, sort_keys=True, indent=4, separators=(',', ': '))
    for d in outBuff:
        print 'Total:%lf'%d['elapse']
        print '%s'%d['start_line']
        print '%s'%d['end_line']
        if not d.has_key('subs'):
            print
            continue
        for s in d['subs']:
            print '\t%lf'%s['elapse']
            print '\t\t%s'%s['start_line']
            print '\t\t%s'%s['end_line']
        print


    wb = Workbook()
    ws = wb.active
    ws.title = 'summary'
    # ws.append(['source file',fn])
    baseLineNumber = 1
    curRowNumber=baseLineNumber;
    sln=-1; # line number
    ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

    for i in range(len(outBuff)):
        sln=curRowNumber;
        d = outBuff[i]
        # print '%s' % d['start_line']
        content = ILLEGAL_CHARACTERS_RE.sub(r'???', d['start_line'])
        ws.append([d['elapse'], content])
        curRowNumber += 1;
        if not d.has_key('subs'):
            ws.append([])
            curRowNumber += 1;
            continue
        for j in range(len(d['subs'])):
            s=d['subs'][j]
            sub_sln=curRowNumber;
            # print '\t%s'%s['start_line']
            subcont_s = ILLEGAL_CHARACTERS_RE.sub(r'???', s['start_line'])
            ws.append(['',s['elapse'],subcont_s])
            subcont_e = ILLEGAL_CHARACTERS_RE.sub(r'???', s['end_line'])
            ws.append(['','',subcont_e])
            curRowNumber += 2;
            ws.merge_cells('B%d:B%d'%(sub_sln,curRowNumber-1))
            sub_cell = ws.cell(row=sub_sln, column=2)
            sub_cell.alignment = Alignment(horizontal='center', vertical='center')
        content_e = ILLEGAL_CHARACTERS_RE.sub(r'???', d['end_line'])
        ws.append(['', content_e])
        ws.append([''])
        curRowNumber+=2
        ws.merge_cells('A%d:A%d'%(sln, curRowNumber-2))
        cell = ws.cell(row=sln, column=1)
        cell.alignment = Alignment(horizontal='center', vertical='center')

    wb.save('AEStation_elapse.xlsx')
    pass



def littleTest(p):

#    l = json.load(file('configure.json', 'r'))
#    p.remove(A[0])
    t=['u1','u2','u3']
    p['u']=t
#    print

import sys
if __name__=='__main__':
#    list=['a','b','c']
#    dic={'a':'av','b':'bv','c':'bv'}
#    littleTest(dic)
#    print dic
    args = sys.argv[:]
    if len(args)<2:
        print 'need parameter [log file name]'
        sys.exit()
    findOutElapse(args[1])
#    print isMatch('.*INFO server.mm run.*', '2017-05-21 00:00:56:257 28167  INFO server.mm run 132 -')
#    t1 = secondsFromDateTimeString('2017-05-21 05:37:01:238')
#    t2 = secondsFromDateTimeString('2017-05-21 05:37:01:338')
#    print t1, t2, t2-t1
    pass
